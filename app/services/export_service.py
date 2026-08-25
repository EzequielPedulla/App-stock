"""Servicio de exportación de reportes a PDF y Excel."""

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate,
    BaseDocTemplate,
    PageTemplate,
    Frame,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
    HRFlowable,
    Image as RLImage
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
import io

from ..utils import formatear_fecha_hora


class ExportService:
    """Servicio para exportar reportes a diferentes formatos."""

    def __init__(self) -> None:
        """Inicializa el servicio de exportación."""
        self.output_dir = Path("reportes")
        self.output_dir.mkdir(exist_ok=True)

    def export_sales_to_excel(
        self,
        ventas: list[dict[str, Any]],
        productos_vendidos: list[dict[str, Any]]
    ) -> str:
        """
        Exporta el historial de ventas a Excel.

        Args:
            ventas: Lista de ventas con sus datos
            productos_vendidos: Lista de productos vendidos con estadísticas

        Returns:
            str: Ruta del archivo generado
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = self.output_dir / f"ventas_{timestamp}.xlsx"

        wb = Workbook()

        # Hoja 1: Historial de Ventas
        ws_ventas = wb.active
        ws_ventas.title = "Historial de Ventas"

        # Encabezado
        headers = ["ID", "Fecha", "Total", "Pagado", "Cambio", "Estado"]
        ws_ventas.append(headers)

        # Estilo del encabezado
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws_ventas[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Datos de ventas
        for venta in ventas:
            fecha_str = formatear_fecha_hora(venta['date'])

            # Determinar el estado de la venta
            estado = venta.get('status', 'active')
            estado_texto = "ACTIVA" if estado == 'active' else "ANULADA"

            ws_ventas.append([
                venta['id'],
                fecha_str,
                float(venta['total']),
                float(venta['paid']),
                float(venta['change']),
                estado_texto
            ])

        # Formatear montos y estado
        for row in range(2, len(ventas) + 2):
            # Formatear montos
            for col in [3, 4, 5]:  # Total, Pagado, Cambio
                cell = ws_ventas.cell(row=row, column=col)
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right")

            # Formatear estado
            estado_cell = ws_ventas.cell(
                row=row, column=6)  # Columna F (Estado)
            estado_cell.alignment = Alignment(horizontal="center")

            # Si la venta está anulada, aplicar formato rojo
            if estado_cell.value == "ANULADA":
                estado_cell.fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )
                estado_cell.font = Font(bold=True, color="CC0000")

        # Ajustar anchos de columna
        ws_ventas.column_dimensions['A'].width = 10
        ws_ventas.column_dimensions['B'].width = 20
        ws_ventas.column_dimensions['C'].width = 15
        ws_ventas.column_dimensions['D'].width = 15
        ws_ventas.column_dimensions['E'].width = 15
        ws_ventas.column_dimensions['F'].width = 12

        # Hoja 2: Productos Más Vendidos
        ws_productos = wb.create_sheet("Productos Vendidos")

        # Encabezado
        headers_prod = ["Producto", "Cantidad Vendida", "Monto Total"]
        ws_productos.append(headers_prod)

        for cell in ws_productos[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Datos de productos
        for producto in productos_vendidos:
            ws_productos.append([
                producto['producto'],
                int(producto['cantidad_vendida']),
                float(producto['monto_total'])
            ])

        # Formatear
        for row in range(2, len(productos_vendidos) + 2):
            ws_productos.cell(row=row, column=2).alignment = Alignment(
                horizontal="center")
            cell = ws_productos.cell(row=row, column=3)
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal="right")

        # Ajustar anchos
        ws_productos.column_dimensions['A'].width = 35
        ws_productos.column_dimensions['B'].width = 20
        ws_productos.column_dimensions['C'].width = 20

        # Agregar bordes a todas las celdas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for ws in [ws_ventas, ws_productos]:
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border

        wb.save(filename)
        return str(filename)

    def export_inventory_to_excel(self, productos: list[dict[str, Any]]) -> str:
        """
        Exporta el inventario de productos a Excel.

        Args:
            productos: Lista de productos con sus datos

        Returns:
            str: Ruta del archivo generado
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = self.output_dir / f"inventario_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        # Encabezado
        headers = ["ID", "Código de Barras", "Nombre", "Precio", "Stock"]
        ws.append(headers)

        # Estilo del encabezado
        header_fill = PatternFill(
            start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Datos
        for producto in productos:
            ws.append([
                producto['id'],
                producto['barcode'],
                producto['name'],
                float(producto['price']),
                producto['stock']
            ])

        # Formatear
        for row in range(2, len(productos) + 2):
            # Precio
            cell = ws.cell(row=row, column=4)
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal="right")

            # Stock
            stock_cell = ws.cell(row=row, column=5)
            stock_cell.alignment = Alignment(horizontal="center")

            # Resaltar stock bajo (< 10)
            if stock_cell.value and stock_cell.value < 10:
                stock_cell.fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                stock_cell.font = Font(bold=True, color="CC0000")

        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12

        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        wb.save(filename)
        return str(filename)

    def export_sales_report_to_pdf(
        self,
        total_ventas: float,
        ultima_venta: float,
        ventas: list[dict[str, Any]],
        productos_vendidos: list[dict[str, Any]]
    ) -> str:
        """
        Exporta un reporte completo de ventas a PDF con gráficos.

        Args:
            total_ventas: Total acumulado de ventas
            ultima_venta: Monto de la última venta
            ventas: Lista de ventas recientes
            productos_vendidos: Lista de productos más vendidos

        Returns:
            str: Ruta del archivo generado
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = self.output_dir / f"reporte_ventas_{timestamp}.pdf"

        doc = SimpleDocTemplate(
            str(filename),
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )

        story = []
        styles = getSampleStyleSheet()

        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e88e5'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        )

        # Título
        story.append(Paragraph("Reporte de Ventas", title_style))
        story.append(Paragraph(
            f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles['Normal']
        ))
        story.append(Spacer(1, 0.3 * inch))

        # Resumen de ventas
        story.append(Paragraph("Resumen General", subtitle_style))

        resumen_data = [
            ['Total de Ventas:', f'${total_ventas:,.2f}'],
            ['Última Venta:', f'${ultima_venta:,.2f}'],
            ['Número de Ventas:', str(len(ventas))]
        ]

        resumen_table = Table(resumen_data, colWidths=[3 * inch, 2 * inch])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
            ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#d5dbdb')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7'))
        ]))

        story.append(resumen_table)
        story.append(Spacer(1, 0.4 * inch))

        # Gráfico de productos más vendidos
        if productos_vendidos:
            story.append(Paragraph("Productos Más Vendidos", subtitle_style))

            # Generar gráfico
            img_buffer = self._create_products_chart(productos_vendidos)
            if img_buffer:
                img = RLImage(img_buffer, width=5 * inch, height=3 * inch)
                story.append(img)
                story.append(Spacer(1, 0.3 * inch))

            # Tabla de productos
            prod_data = [['Producto', 'Cantidad', 'Monto Total']]
            for prod in productos_vendidos[:10]:
                prod_data.append([
                    prod['producto'][:30],
                    str(int(prod['cantidad_vendida'])),
                    f"${float(prod['monto_total']):,.2f}"
                ])

            prod_table = Table(
                prod_data, colWidths=[3.5 * inch, 1 * inch, 1.5 * inch])
            prod_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#26a69a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1),
                 [colors.white, colors.HexColor('#f5f5f5')])
            ]))

            story.append(prod_table)

        # Nueva página para historial
        story.append(PageBreak())
        story.append(Paragraph("Historial de Ventas", subtitle_style))
        story.append(Spacer(1, 0.2 * inch))

        # Tabla de ventas (mostrar las últimas 20)
        ventas_data = [['ID', 'Fecha', 'Total']]
        for venta in ventas[:20]:
            fecha_str = formatear_fecha_hora(venta['date'])

            ventas_data.append([
                str(venta['id']),
                fecha_str,
                f"${float(venta['total']):,.2f}"
            ])

        ventas_table = Table(
            ventas_data, colWidths=[0.8 * inch, 2.5 * inch, 1.5 * inch])
        ventas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e88e5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#e3f2fd')])
        ]))

        story.append(ventas_table)

        # Pie de página
        story.append(Spacer(1, 0.5 * inch))
        footer_text = f"Total de ventas mostradas: {len(ventas[:20])} de {len(ventas)}"
        story.append(Paragraph(footer_text, styles['Italic']))

        doc.build(story)
        return str(filename)

    def export_sale_ticket_to_pdf(
        self,
        sale_id: int,
        sale_date: str,
        payment_method: str,
        sale_total: float,
        sale_paid: float,
        sale_change: float,
        details: list[dict[str, Any]]
    ) -> str:
        """
        Exporta un ticket/recibo de venta individual a PDF, pensado para
        imprimirse en una impresora térmica de 58mm de rollo (48-50mm de
        ancho útil de impresión).

        Args:
            sale_id: ID de la venta (solo se usa para nombrar el archivo;
                no se imprime en el ticket, no le sirve al cliente)
            sale_date: Fecha de la venta (ya formateada, con hora)
            payment_method: Método de pago ('efectivo', 'mixto', etc.)
            sale_total: Total de la venta
            sale_paid: Monto pagado
            sale_change: Cambio entregado
            details: Detalles de los productos vendidos

        Returns:
            str: Ruta del archivo generado
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = self.output_dir / f"ticket_venta_{sale_id}_{timestamp}.pdf"

        ancho_pagina = 58 * mm
        margen = 3 * mm
        margen_vertical = 3 * mm
        ancho_util = ancho_pagina - (2 * margen)

        story = []
        styles = getSampleStyleSheet()

        titulo_style = ParagraphStyle(
            'TicketTitulo', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=12, alignment=TA_CENTER,
            leading=14)
        subtitulo_style = ParagraphStyle(
            'TicketSubtitulo', parent=styles['Normal'],
            fontName='Helvetica', fontSize=7.5, alignment=TA_CENTER,
            leading=9, textColor=colors.HexColor('#555555'))
        normal_style = ParagraphStyle(
            'TicketNormal', parent=styles['Normal'],
            fontName='Helvetica', fontSize=8, leading=10)
        celda_style = ParagraphStyle(
            'TicketCelda', parent=normal_style, fontSize=6.8, leading=8.2)
        celda_header_style = ParagraphStyle(
            'TicketCeldaHeader', parent=celda_style,
            fontName='Helvetica-Bold')
        celda_header_derecha_style = ParagraphStyle(
            'TicketCeldaHeaderDerecha', parent=celda_header_style,
            alignment=TA_RIGHT)
        celda_derecha_style = ParagraphStyle(
            'TicketCeldaDerecha', parent=celda_style, alignment=TA_RIGHT)
        derecha_style = ParagraphStyle(
            'TicketDerecha', parent=normal_style, alignment=TA_RIGHT)
        total_style = ParagraphStyle(
            'TicketTotal', parent=normal_style,
            fontName='Helvetica-Bold', fontSize=11, alignment=TA_RIGHT)
        pie_style = ParagraphStyle(
            'TicketPie', parent=styles['Normal'],
            fontName='Helvetica-Oblique', fontSize=6.5,
            alignment=TA_CENTER, textColor=colors.HexColor('#777777'),
            leading=8)

        metodo_labels = {
            'efectivo': 'Efectivo',
            'transferencia': 'Transferencia',
            'posnet': 'Posnet',
            'fiado': 'Fiado (pendiente de cobro)',
            'mixto': 'Efectivo + Transferencia',
        }

        # Líneas separadoras como filete real (no texto "="/"-"): se arman
        # con Spacer explícito antes/después en vez de spaceBefore/
        # spaceAfter, para que el alto que ocupan se pueda medir con
        # wrap() más abajo (ver por qué, en el comentario de alto_pagina).
        def _agregar_linea(fuerte: bool) -> None:
            story.append(Spacer(1, 1.5 * mm))
            story.append(HRFlowable(
                width="100%",
                thickness=0.8 if fuerte else 0.4,
                color=colors.black if fuerte else colors.HexColor('#999999'),
                spaceBefore=0, spaceAfter=0
            ))
            story.append(Spacer(1, 1.5 * mm))

        # Encabezado
        story.append(Paragraph("App-Stock", titulo_style))
        story.append(Paragraph("Ticket de venta", subtitulo_style))
        _agregar_linea(fuerte=True)

        # Datos de la venta (el N° de venta/del día es interno, no se
        # muestra en el ticket: no le sirve al cliente)
        story.append(Paragraph(f"Fecha: {sale_date}", normal_style))
        story.append(Paragraph(
            f"Pago: {metodo_labels.get(payment_method, 'Efectivo')}",
            normal_style))
        _agregar_linea(fuerte=False)

        # Tabla de productos: todas las celdas usan Paragraph (no strings
        # sueltos), así un nombre largo o un monto grande se acomodan en
        # una segunda línea en vez de pisar la columna de al lado.
        col_producto = ancho_util * 0.28
        col_cant = ancho_util * 0.15
        col_precio = ancho_util * 0.285
        col_subtotal = ancho_util * 0.285

        products_data = [[
            Paragraph('Producto', celda_header_style),
            Paragraph('Ud.', celda_header_derecha_style),
            Paragraph('P. Unit.', celda_header_derecha_style),
            Paragraph('Subtot.', celda_header_derecha_style),
        ]]
        for detail in details:
            products_data.append([
                Paragraph(str(detail['producto']), celda_style),
                Paragraph(str(detail['cantidad']), celda_derecha_style),
                Paragraph(f"${detail['precio']:.2f}", celda_derecha_style),
                Paragraph(f"${detail['subtotal']:.2f}", celda_derecha_style),
            ])

        products_table = Table(
            products_data,
            colWidths=[col_producto, col_cant, col_precio, col_subtotal]
        )
        products_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LINEBELOW', (0, 0), (-1, 0), 0.6, colors.black),
            ('TOPPADDING', (0, 0), (-1, -1), 1.5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            # Un pequeño espacio entre columnas para que los montos no
            # queden pegados al de al lado.
            ('RIGHTPADDING', (0, 0), (2, -1), 4),
            ('LEFTPADDING', (1, 0), (3, -1), 3),
        ]))
        story.append(products_table)
        _agregar_linea(fuerte=False)

        # Totales: apilados y alineados a la derecha, como en un ticket real.
        story.append(Paragraph(f"TOTAL: ${sale_total:.2f}", total_style))
        if payment_method == 'fiado':
            story.append(Paragraph(
                "Sin cobrar (fiado)", derecha_style))
        else:
            story.append(Paragraph(f"Pagado: ${sale_paid:.2f}", derecha_style))
            if sale_change > 0.005:
                story.append(Paragraph(
                    f"Vuelto: ${sale_change:.2f}", derecha_style))

        _agregar_linea(fuerte=True)

        # Pie de página
        story.append(Paragraph("¡Gracias por su compra!", subtitulo_style))
        story.append(Paragraph(
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            pie_style
        ))

        # La impresora es de rollo continuo, pero reportlab necesita un
        # tamaño de página fijo de antemano. En vez de adivinar el alto
        # (un nombre largo o un monto grande pueden ocupar más líneas de
        # lo esperado y desbordar a una segunda página), se mide el alto
        # real que ocupa cada elemento ya armado y se arma la página a
        # medida: así el ticket entra siempre en una sola hoja.
        alto_contenido = 0.0
        for flowable in story:
            _, alto = flowable.wrap(ancho_util, 5000)
            alto_contenido += alto
        alto_pagina = alto_contenido + (2 * margen_vertical) + (2 * mm)

        # No se usa SimpleDocTemplate: su Frame interno agrega 6pt de
        # padding propio a cada lado (además del margen ya puesto acá),
        # lo que desajustaba el ancho/alto real respecto de lo medido
        # arriba. Con un Frame armado a mano, el padding es el margen
        # elegido y nada más, así la medición de arriba es exacta.
        doc = BaseDocTemplate(
            str(filename),
            pagesize=(ancho_pagina, alto_pagina),
            leftMargin=0, rightMargin=0, topMargin=0, bottomMargin=0
        )
        frame = Frame(
            0, 0, ancho_pagina, alto_pagina,
            leftPadding=margen, rightPadding=margen,
            topPadding=margen_vertical, bottomPadding=margen_vertical,
            id='ticket'
        )
        doc.addPageTemplates([PageTemplate(id='ticket', frames=[frame])])
        doc.build(story)
        return str(filename)

    def _create_products_chart(
        self, productos_vendidos: list[dict[str, Any]]
    ) -> io.BytesIO | None:
        """
        Crea un gráfico de barras de productos más vendidos.

        Args:
            productos_vendidos: Lista de productos con cantidades vendidas

        Returns:
            BytesIO: Buffer con la imagen del gráfico, o None si hay error
        """
        try:
            top_productos = productos_vendidos[:5]
            nombres = [p['producto'][:20] for p in top_productos]
            cantidades = [int(p['cantidad_vendida']) for p in top_productos]

            fig = Figure(figsize=(6, 3.5), dpi=100)
            ax = fig.add_subplot(111)

            colores = ['#1e88e5', '#26a69a', '#66bb6a', '#ffa726', '#ef5350']

            bars = ax.bar(nombres, cantidades, color=colores[:len(nombres)],
                          width=0.65, edgecolor='white', linewidth=1.5, alpha=0.9)

            ax.yaxis.grid(True, linestyle='--', alpha=0.3, color='gray')
            ax.set_axisbelow(True)

            ax.set_ylabel('Unidades Vendidas', fontsize=10, weight='bold')
            ax.set_xlabel('Productos', fontsize=10, weight='bold')
            ax.set_title('Top 5 Productos Más Vendidos',
                         fontsize=12, weight='bold', pad=10)

            plt.setp(ax.xaxis.get_majorticklabels(),
                     rotation=35, ha='right', fontsize=9)

            for bar, cantidad in zip(bars, cantidades):
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2, height,
                        f'{cantidad}',
                        ha='center', va='bottom', fontsize=10, weight='bold')

            ax.set_ylim(0, max(cantidades) * 1.15)

            fig.tight_layout()

            buffer = io.BytesIO()
            fig.savefig(buffer, format='png', bbox_inches='tight', dpi=100)
            buffer.seek(0)
            plt.close(fig)

            return buffer

        except Exception as e:
            print(f"Error al crear gráfico: {e}")
            return None
