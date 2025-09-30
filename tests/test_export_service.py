"""Tests para el servicio de exportación."""

from pathlib import Path
from typing import TYPE_CHECKING
import pytest
from app.services.export_service import ExportService

if TYPE_CHECKING:
    from _pytest.fixtures import FixtureRequest
    from pytest_mock.plugin import MockerFixture


@pytest.fixture
def export_service() -> ExportService:
    """
    Fixture que proporciona una instancia del servicio de exportación.

    Returns:
        ExportService: Instancia del servicio de exportación
    """
    service = ExportService()
    # Crear directorio temporal para tests
    service.output_dir = Path("test_reportes")
    service.output_dir.mkdir(exist_ok=True)
    yield service
    # Limpiar después de los tests
    if service.output_dir.exists():
        for file in service.output_dir.iterdir():
            file.unlink()
        service.output_dir.rmdir()


@pytest.fixture
def sample_ventas() -> list[dict[str, any]]:
    """
    Fixture que proporciona datos de ejemplo de ventas.

    Returns:
        list: Lista de ventas de ejemplo
    """
    return [
        {
            'id': 1,
            'date': '2024-01-15 10:30:00',
            'total': 150.50,
            'paid': 200.00,
            'change': 49.50
        },
        {
            'id': 2,
            'date': '2024-01-15 14:20:00',
            'total': 75.25,
            'paid': 100.00,
            'change': 24.75
        },
        {
            'id': 3,
            'date': '2024-01-16 09:15:00',
            'total': 200.00,
            'paid': 200.00,
            'change': 0.00
        }
    ]


@pytest.fixture
def sample_productos_vendidos() -> list[dict[str, any]]:
    """
    Fixture que proporciona datos de ejemplo de productos vendidos.

    Returns:
        list: Lista de productos vendidos de ejemplo
    """
    return [
        {
            'producto': 'Caldo',
            'cantidad_vendida': 50,
            'monto_total': 75.00
        },
        {
            'producto': 'Pan',
            'cantidad_vendida': 40,
            'monto_total': 80.00
        },
        {
            'producto': 'Jugo',
            'cantidad_vendida': 30,
            'monto_total': 30.00
        },
        {
            'producto': 'Galletas',
            'cantidad_vendida': 25,
            'monto_total': 75.00
        }
    ]


@pytest.fixture
def sample_productos() -> list[dict[str, any]]:
    """
    Fixture que proporciona datos de ejemplo de productos.

    Returns:
        list: Lista de productos de ejemplo
    """
    return [
        {
            'id': 1,
            'barcode': '123456789012',
            'name': 'Caldo',
            'price': 1.50,
            'stock': 40
        },
        {
            'id': 2,
            'barcode': '987654322108',
            'name': 'Pan',
            'price': 2.00,
            'stock': 25
        },
        {
            'id': 3,
            'barcode': '456789012345',
            'name': 'Jugo',
            'price': 1.00,
            'stock': 5  # Stock bajo
        }
    ]


@pytest.fixture
def sample_sale_details() -> list[dict[str, any]]:
    """
    Fixture que proporciona detalles de una venta de ejemplo.

    Returns:
        list: Lista de detalles de venta
    """
    return [
        {
            'producto': 'Caldo',
            'cantidad': 3,
            'precio': 1.50,
            'subtotal': 4.50
        },
        {
            'producto': 'Pan',
            'cantidad': 2,
            'precio': 2.00,
            'subtotal': 4.00
        }
    ]


class TestExportService:
    """Tests para ExportService."""

    def test_export_service_initialization(self, export_service: ExportService) -> None:
        """
        Test de inicialización del servicio.

        Args:
            export_service: Fixture del servicio de exportación
        """
        assert export_service is not None
        assert export_service.output_dir.exists()

    def test_export_sales_to_excel(
        self,
        export_service: ExportService,
        sample_ventas: list[dict[str, any]],
        sample_productos_vendidos: list[dict[str, any]]
    ) -> None:
        """
        Test de exportación de ventas a Excel.

        Args:
            export_service: Fixture del servicio de exportación
            sample_ventas: Fixture con ventas de ejemplo
            sample_productos_vendidos: Fixture con productos vendidos
        """
        # Exportar
        filename = export_service.export_sales_to_excel(
            sample_ventas, sample_productos_vendidos
        )

        # Verificar que el archivo fue creado
        assert Path(filename).exists()
        assert filename.endswith('.xlsx')
        assert 'ventas_' in filename

        # Verificar que es un archivo válido de Excel
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        assert 'Historial de Ventas' in wb.sheetnames
        assert 'Productos Vendidos' in wb.sheetnames

        # Verificar datos en hoja de ventas
        ws_ventas = wb['Historial de Ventas']
        assert ws_ventas['A1'].value == 'ID'
        assert ws_ventas['B1'].value == 'Fecha'
        assert ws_ventas['C1'].value == 'Total'

        # Verificar que hay datos (3 ventas + 1 encabezado)
        assert ws_ventas.max_row == 4

        # Verificar datos en hoja de productos
        ws_productos = wb['Productos Vendidos']
        assert ws_productos['A1'].value == 'Producto'
        assert ws_productos['B1'].value == 'Cantidad Vendida'
        assert ws_productos['C1'].value == 'Monto Total'

        # Verificar que hay datos (4 productos + 1 encabezado)
        assert ws_productos.max_row == 5

        wb.close()

    def test_export_inventory_to_excel(
        self,
        export_service: ExportService,
        sample_productos: list[dict[str, any]]
    ) -> None:
        """
        Test de exportación de inventario a Excel.

        Args:
            export_service: Fixture del servicio de exportación
            sample_productos: Fixture con productos de ejemplo
        """
        # Exportar
        filename = export_service.export_inventory_to_excel(sample_productos)

        # Verificar que el archivo fue creado
        assert Path(filename).exists()
        assert filename.endswith('.xlsx')
        assert 'inventario_' in filename

        # Verificar contenido
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        ws = wb.active

        assert ws['A1'].value == 'ID'
        assert ws['B1'].value == 'Código de Barras'
        assert ws['C1'].value == 'Nombre'
        assert ws['D1'].value == 'Precio'
        assert ws['E1'].value == 'Stock'

        # Verificar que hay datos (3 productos + 1 encabezado)
        assert ws.max_row == 4

        # Verificar que el stock bajo está resaltado
        stock_bajo_cell = ws['E4']  # Jugo tiene stock 5
        assert stock_bajo_cell.value == 5
        # Verificar que tiene formato de advertencia (puede ser FFFFCCCC o 00FFCCCC)
        assert 'FFCCCC' in stock_bajo_cell.fill.start_color.rgb

        wb.close()

    def test_export_sales_report_to_pdf(
        self,
        export_service: ExportService,
        sample_ventas: list[dict[str, any]],
        sample_productos_vendidos: list[dict[str, any]]
    ) -> None:
        """
        Test de exportación de reporte de ventas a PDF.

        Args:
            export_service: Fixture del servicio de exportación
            sample_ventas: Fixture con ventas de ejemplo
            sample_productos_vendidos: Fixture con productos vendidos
        """
        # Exportar
        filename = export_service.export_sales_report_to_pdf(
            total_ventas=425.75,
            ultima_venta=200.00,
            ventas=sample_ventas,
            productos_vendidos=sample_productos_vendidos
        )

        # Verificar que el archivo fue creado
        assert Path(filename).exists()
        assert filename.endswith('.pdf')
        assert 'reporte_ventas_' in filename

        # Verificar que es un PDF válido (tiene el header)
        with open(filename, 'rb') as f:
            header = f.read(5)
            assert header == b'%PDF-'

    def test_export_sale_ticket_to_pdf(
        self,
        export_service: ExportService,
        sample_sale_details: list[dict[str, any]]
    ) -> None:
        """
        Test de exportación de ticket de venta a PDF.

        Args:
            export_service: Fixture del servicio de exportación
            sample_sale_details: Fixture con detalles de venta
        """
        # Exportar
        filename = export_service.export_sale_ticket_to_pdf(
            sale_id=1,
            sale_date='2024-01-15 10:30:00',
            sale_total=8.50,
            sale_paid=10.00,
            sale_change=1.50,
            details=sample_sale_details
        )

        # Verificar que el archivo fue creado
        assert Path(filename).exists()
        assert filename.endswith('.pdf')
        assert 'ticket_venta_1_' in filename

        # Verificar que es un PDF válido
        with open(filename, 'rb') as f:
            header = f.read(5)
            assert header == b'%PDF-'

    def test_export_sales_to_excel_empty_data(
        self,
        export_service: ExportService
    ) -> None:
        """
        Test de exportación con datos vacíos.

        Args:
            export_service: Fixture del servicio de exportación
        """
        # Exportar con listas vacías
        filename = export_service.export_sales_to_excel([], [])

        # Verificar que el archivo fue creado igual
        assert Path(filename).exists()

        from openpyxl import load_workbook
        wb = load_workbook(filename)

        # Solo debe tener encabezados
        ws_ventas = wb['Historial de Ventas']
        assert ws_ventas.max_row == 1  # Solo encabezado

        wb.close()

    def test_create_products_chart_success(
        self,
        export_service: ExportService,
        sample_productos_vendidos: list[dict[str, any]]
    ) -> None:
        """
        Test de creación de gráfico de productos.

        Args:
            export_service: Fixture del servicio de exportación
            sample_productos_vendidos: Fixture con productos vendidos
        """
        # Crear gráfico
        buffer = export_service._create_products_chart(
            sample_productos_vendidos)

        # Verificar que se creó el buffer
        assert buffer is not None

        # Verificar que es una imagen PNG válida
        buffer.seek(0)
        header = buffer.read(8)
        assert header == b'\x89PNG\r\n\x1a\n'

        # Verificar que tiene contenido después del header
        buffer.seek(0)
        content = buffer.read()
        assert len(content) > 100  # Una imagen PNG debe tener más de 100 bytes

    def test_create_products_chart_empty_data(
        self,
        export_service: ExportService
    ) -> None:
        """
        Test de creación de gráfico con datos vacíos.

        Args:
            export_service: Fixture del servicio de exportación
        """
        # Intentar crear gráfico con lista vacía
        buffer = export_service._create_products_chart([])

        # Debe retornar None o un buffer vacío
        # (dependiendo de la implementación)
        # En este caso, debería fallar gracefully
        assert buffer is None or buffer.tell() == 0

    def test_export_sales_to_excel_with_special_characters(
        self,
        export_service: ExportService
    ) -> None:
        """
        Test de exportación con caracteres especiales.

        Args:
            export_service: Fixture del servicio de exportación
        """
        ventas_especiales = [
            {
                'id': 1,
                'date': '2024-01-15 10:30:00',
                'total': 100.00,
                'paid': 100.00,
                'change': 0.00
            }
        ]

        productos_especiales = [
            {
                'producto': 'Café con Leche & Azúcar',
                'cantidad_vendida': 10,
                'monto_total': 50.00
            }
        ]

        # Exportar
        filename = export_service.export_sales_to_excel(
            ventas_especiales, productos_especiales
        )

        # Verificar que el archivo fue creado
        assert Path(filename).exists()

        from openpyxl import load_workbook
        wb = load_workbook(filename)
        ws_productos = wb['Productos Vendidos']

        # Verificar que los caracteres especiales se guardaron correctamente
        assert 'Café con Leche & Azúcar' in ws_productos['A2'].value

        wb.close()

    def test_export_large_dataset(
        self,
        export_service: ExportService
    ) -> None:
        """
        Test de exportación con dataset grande.

        Args:
            export_service: Fixture del servicio de exportación
        """
        # Crear dataset grande (100 ventas)
        large_ventas = [
            {
                'id': i,
                'date': f'2024-01-15 10:{i:02d}:00',
                'total': 100.00 + i,
                'paid': 150.00,
                'change': 50.00 - i
            }
            for i in range(100)
        ]

        large_productos = [
            {
                'producto': f'Producto {i}',
                'cantidad_vendida': 10 + i,
                'monto_total': 50.00 + i
            }
            for i in range(50)
        ]

        # Exportar
        filename = export_service.export_sales_to_excel(
            large_ventas, large_productos
        )

        # Verificar que el archivo fue creado
        assert Path(filename).exists()

        from openpyxl import load_workbook
        wb = load_workbook(filename)

        # Verificar cantidad de filas
        ws_ventas = wb['Historial de Ventas']
        assert ws_ventas.max_row == 101  # 100 ventas + 1 encabezado

        ws_productos = wb['Productos Vendidos']
        assert ws_productos.max_row == 51  # 50 productos + 1 encabezado

        wb.close()
